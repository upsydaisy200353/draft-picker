"""检测 Excel 中标蓝的行/单元格"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX

path = Path(r'c:\Users\32828\Desktop\白菜杯S46海斗杯报名表（收集结果）.xlsx')
wb = load_workbook(path, data_only=True)
ws = wb.active

def color_info(cell):
    fill = cell.fill
    if not fill or fill.fill_type != 'solid':
        return None
    fg = fill.fgColor
    if not fg:
        return None
    if fg.type == 'rgb' and fg.rgb:
        return fg.rgb
    if fg.type == 'indexed' and fg.indexed is not None:
        return f'indexed:{fg.indexed}'
    if fg.type == 'theme':
        return f'theme:{fg.theme}'
    return str(fg)

print('Sheet:', ws.title)
print('Rows:', ws.max_row, 'Cols:', ws.max_column)
print()

# header
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print('Headers:', headers)
print()

blue_rows = []
for r in range(2, ws.max_row + 1):
    row_colors = []
    row_vals = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c)
        row_vals.append(cell.value)
        ci = color_info(cell)
        if ci:
            row_colors.append((c, ci))
    # print all non-default colors for inspection
    if row_colors:
        print(f'Row {r}: colors={row_colors}')
        print(f'  values={row_vals}')
        blue_rows.append((r, row_vals, row_colors))

print()
print('Rows with any fill color:', len(blue_rows))
