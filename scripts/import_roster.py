"""从报名表 xlsx 生成 draft-picker 的 config.json（队长 = 行标蓝）"""
import json
import re
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(r'c:\Users\32828\Desktop\白菜杯S46海斗杯报名表（收集结果）.xlsx')
OUT = Path(__file__).resolve().parent.parent / 'data' / 'config.json'

SKILL_STRENGTH = {
    '能躺': 1,
    '能送': 2,
    '能C能送': 3,
    '包C': 4,
    '一打九': 5,
}

BLUE_RGB_PREFIXES = ('FF8CDDFA', 'FF9BC2E6', 'FFBDD7EE', 'FF4472C4', 'FF5B9BD5', 'FF00B0F0')


def slugify(text: str) -> str:
    text = str(text).strip().lower()
    text = re.sub(r'[^\w\u4e00-\u9fff]+', '_', text, flags=re.UNICODE)
    text = re.sub(r'_+', '_', text).strip('_')
    return text[:24] or 'player'


def cell_rgb(cell):
    fill = cell.fill
    if not fill or fill.fill_type != 'solid':
        return None
    fg = fill.fgColor
    if fg and fg.type == 'rgb' and fg.rgb:
        return fg.rgb.upper()
    return None


def is_blue_cell(cell):
    rgb = cell_rgb(cell)
    if not rgb:
        return False
    if rgb in BLUE_RGB_PREFIXES:
        return True
    if len(rgb) == 8:
        try:
            r = int(rgb[2:4], 16)
            g = int(rgb[4:6], 16)
            b = int(rgb[6:8], 16)
            return b > 160 and b > r and b > g - 40
        except ValueError:
            return False
    return False


def row_is_blue(ws, row_idx):
    for col in range(1, ws.max_column + 1):
        if is_blue_cell(ws.cell(row_idx, col)):
            return True
    return False


def main():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active

    captains_raw = []
    players_raw = []

    for r in range(2, ws.max_row + 1):
        nickname = str(ws.cell(r, 3).value or '').strip()
        game_id = str(ws.cell(r, 4).value or '').strip()
        skill = str(ws.cell(r, 5).value or '').strip()
        display = nickname if nickname and nickname != 'None' else game_id

        entry = {
            'display': display,
            'nickname': nickname,
            'game_id': game_id,
            'skill': skill,
            'strength_score': SKILL_STRENGTH.get(skill, 3),
        }

        if row_is_blue(ws, r):
            captains_raw.append(entry)
        else:
            players_raw.append(entry)

    captains_raw.sort(key=lambda x: (x['strength_score'], x['display']))

    used_usernames = {'admin'}
    captains = []
    for i, c in enumerate(captains_raw, 1):
        username = f'captain{i}'
        captains.append({
            'id': f'c{i}',
            'name': c['display'],
            'username': username,
            'password': f'hd{i:02d}',
            'strength': i,
            'skill': c['skill'],
            'game_id': c['game_id'],
        })

    players = []
    for i, p in enumerate(players_raw, 1):
        players.append({
            'id': f'p{i}',
            'name': p['display'],
            'username': f'player{i}',
            'password': f'mp{i:02d}',
            'game_id': p['game_id'],
            'skill': p['skill'],
        })

    config = {
        'admin': {'username': 'admin', 'password': 'admin123'},
        'event': '白菜杯S46海斗杯',
        'captain_rule': '报名表行标蓝为队长',
        'captains': captains,
        'players': players,
    }

    OUT.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'已写入 {OUT}')
    print(f'队长 {len(captains)} 人，选手池 {len(players)} 人')
    print()
    print('队长账号（实力弱→强，strength 越小越先抽）：')
    for c in captains:
        print(f"  {c['strength']}. {c['name']}（{c['skill']}）→ {c['username']} / {c['password']}")


if __name__ == '__main__':
    main()
